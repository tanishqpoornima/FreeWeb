from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import pyodbc
import numpy as np


app = Flask(__name__)

# Set up Azure SQL Database connection details
server = 'fl.database.windows.net'
database = 'FL'
username = 'admin-fl'
password = '123@Fl-@'
# driver= '{ODBC Driver 17 for SQL Server}'
driver = '{SQL Server}'

# Establish a connection to the database
def create_connection():
    return pyodbc.connect(f'DRIVER={driver};SERVER={server};DATABASE={database};UID={username};PWD={password}')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    name = request.form['name']
    tname = request.form['name']
    key = request.form['key']
    cost = request.form['cost']

    # Here you can add code to store the data in an Excel file or a database


    # Connect to the Azure SQL Database
    conn = create_connection()
    cursor = conn.cursor()
    l = np.array(list(cursor.execute("SELECT * FROM INFORMATION_SCHEMA.TABLES")))
    # print(l[:])
    if name not in l[:]:
        cursor.execute(f"CREATE TABLE {name} (Name varchar(30), SubjectKey varchar(30), Cost int, PRIMARY KEY (SubjectKey));")
        conn.commit()
    try:
        cursor.execute(f"INSERT INTO {name} (Name, SubjectKey, Cost) VALUES (?, ?, ?)", name, key, cost)
        conn.commit()
        save_to_excel(name,key,cost)
    except:
        return render_template('submitted.html', name=name, key=key, cost=cost)
    # Close the database connection
    conn.close()

    # return "Data has been stored in the Azure SQL Database."

    return render_template('submitted.html', name=name, key=key, cost=cost)


def save_to_excel(name,key,cost):


    try:
        wb = load_workbook('data.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(title=name)

    if not ws['A1'].value:
        ws['A1'] = 'Name'
        ws['B1'] = 'Key'
        ws['C1'] = 'Cost'

    row = ws.max_row + 1
    ws.cell(row=row, column=1, value=name)
    ws.cell(row=row, column=2, value=key)
    ws.cell(row=row, column=3, value=cost)

    wb.save('data.xlsx')
    # status_label.config(text="Data saved successfully", fg="green")

# def clear_entries(name.get(),key.get(),cost.get()):
#     nam`e.delete(0, tk.END)
#     age_entry.delete(0, tk.END)
#     email_entry.delete(0, tk.END)

if __name__ == '__main__':
    app.run(debug=True)






